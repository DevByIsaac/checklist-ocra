import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
import datetime 
from modules.analisis_video1 import *
from modules.tiempo_neto2 import *
from modules.factor_recuperacion3 import *
from modules.factor_frecuencia4 import *
from modules.factor_fuerza5 import *
from modules.factor_posturas_movimientos6 import *
from modules.factor_riesgos_adicionales7 import *
from modules.multiplicador_duracion8 import *
from modules.evaluacion_ocra9 import *

def convertir_datafram(data_list):
    # Leer el JSON y convertirlo a un DataFrame de pandas
    df = pd.DataFrame(data_list)

    # Obtener valores para el cálculo de ICKL
    fr = df['Factor de Recuperacion'].mean()  # Promedio de los valores si hay múltiples registros
    ff = df['FF'].mean()
    ffz = df['Factor de Fuerza'].mean()
    fp = df['Factor de Posturas y Movimientos'].mean()
    fc = df['Factor de Riesgos Adicionales'].mean()
    md = df['Multiplicador de Duracion'].mean()

    # Calcular el Índice Check List OCRA (ICKL)
    ickl = (fr + ff + ffz + fp + fc) * md
    path = armar_excel(ickl, df, fr, ff, ffz, fp, fc, md)
    return path
# Determinar Nivel de Riesgo y Acción Recomendada
def determinar_nivel_riesgo(ickl):
    if ickl <= 5:
        return 'Óptimo', 'No se requiere', '≤ 1.5'
    elif 5.1 <= ickl <= 7.5:
        return 'Aceptable', 'No se requiere', '1.6 - 2.2'
    elif 7.6 <= ickl <= 11:
        return 'Incierto', 'Se recomienda un nuevo análisis o mejora del puesto', '2.3 - 3.5'
    elif 11.1 <= ickl <= 14:
        return 'Inaceptable Leve', 'Se recomienda mejora del puesto, supervisión médica y entrenamiento', '3.6 - 4.5'
    elif 14.1 <= 22.5:
        return 'Inaceptable Medio', 'Se recomienda mejora del puesto, supervisión médica y entrenamiento', '4.6 - 9'
    else:
        return 'Inaceptable Alto', 'Se recomienda mejora del puesto, supervisión médica y entrenamiento', '> 9'

def armar_excel(ickl, df, fr, ff, ffz, fp, fc, md):
    nivel_riesgo, accion_recomendada, indice_ocra_equivalente = determinar_nivel_riesgo(ickl)

    # Obtener el nombre del video y la fecha actual
    #nombre_video = df['analisis_video'][0].split('\\')[-1].replace('.mp4', '')
    nombre_video = os.path.basename(df['analisis_video'][0]).replace('.mp4', '')

    fecha_hora = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

    # Construir el nombre del archivo Excel
    archivo_excel = f'analisisOcra_{nombre_video}_{fecha_hora}.xlsx'

    # Definir la ruta de la carpeta 'resultados' dentro de 'static'
    static_folder = 'static'
    resultados_folder = os.path.join(static_folder, 'resultados')

    # Verifica si la carpeta 'uploads' existe, y si no, créala
    uploads_folder = os.path.join(static_folder, 'uploads')
    if not os.path.exists(uploads_folder):
        os.makedirs(uploads_folder)

    # Crear la carpeta 'resultados' si no existe
    if not os.path.exists(resultados_folder):
        os.makedirs(resultados_folder)

    # Definir la ruta del archivo Excel en la carpeta 'resultados'
    excel_path = os.path.join(resultados_folder, archivo_excel)

    # Crear un nuevo libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen OCRA"

    # Escribir encabezados para los factores
    headers_factores = [
        "Factor de Recuperación (FR)", "Factor de Frecuencia (FF)", 
        "Factor de Fuerza (FFz)", "Factor de Posturas y Movimientos (FP)",
        "Factor de Riesgos Adicionales (FC)", "Multiplicador de Duración (MD)"
    ]
    ws.append(headers_factores)

    # Escribir la fila de factores
    factores = [fr, ff, ffz, fp, fc, md]
    ws.append(factores)

    # Escribir encabezados para los resultados del Check List OCRA
    ws.append([])  # Añadir una fila en blanco para separación
    headers_resultados = [
        "Índice Check List OCRA", "Nivel de Riesgo", "Acción Recomendada", "Índice OCRA Equivalente"
    ]
    ws.append(headers_resultados)

    # Escribir la fila de resultados
    resultados = [ickl, nivel_riesgo, accion_recomendada, indice_ocra_equivalente]
    ws.append(resultados)

    # Estilos para la hoja de cálculo
    header_fill = PatternFill(start_color='4C47EA', end_color='4C47EA', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar estilo a los encabezados de factores
    for col_num in range(1, 7):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Aplicar estilo a los encabezados de resultados
    for col_num in range(1, 5):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Aplicar formato a las celdas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment

            # Aplicar formato condicional para los factores en la fila 2
            """ if cell.row == 2:  # Fila de factores
                if isinstance(cell.value, (int, float)):
                    if cell.value > 14:
                        # Color: Rojo claro (FF9999)
                        cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                    elif cell.value > 7.5:
                        # Color: Amarillo claro (FFCC99)
                        cell.fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid') """
            for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = alignment
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 14:
                            # Color: Rojo claro (FF9999)
                            cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                        elif cell.value > 7.5:
                            # Color: Amarillo claro (FFCC99)
                            cell.fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')

            # Aplicar formato condicional para los resultados en la fila 5
            """ if cell.row == 5:  # Fila de resultados
                if isinstance(cell.value, (int, float)):
                    if cell.value > 14:
                        # Color: Rojo claro (FF9999)
                        cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                    elif cell.value > 7.5:
                        # Color: Amarillo claro (FFCC99)
                        cell.fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid') """
            # Aplicar formato condicional para los resultados en la fila 5
            for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = alignment
                    if isinstance(cell.value, (int, float)):
                        if cell.value <= 5:
                            # Color: Verde encendido (9BBB59)
                            cell.fill = PatternFill(start_color='9BBB59', end_color='9BBB59', fill_type='solid')
                        elif 5.1 <= cell.value <= 7.5:
                            # Color: Verde claro (C5E0B4)
                            cell.fill = PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid')
                        elif 7.6 <= cell.value <= 11:
                            # Color: Naranja pastel (F4B084)
                            cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
                        elif 11.1 <= cell.value <= 14:
                            # Color: Vino claro (C65911)
                            cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
                        elif 14.1 <= cell.value <= 22.5:
                            # Color: Rojo pastel (FF9999)
                            cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
                        else:  # > 22.5
                            # Color: Rojo (FF0000)
                            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass    
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar el archivo Excel
    wb.save(excel_path)

    print(f"El Excel se ha creado y guardado en {excel_path}.")
    return excel_path